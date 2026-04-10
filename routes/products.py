import csv
import io
import json
import re
from datetime import datetime, UTC
from pathlib import Path
from typing import Optional
from fastapi import APIRouter, Request, HTTPException, Query
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse
from utils.serializer import clean

router = APIRouter()


@router.get("/")
async def get_products(
    request: Request,
    platform: Optional[str] = None,
    category: Optional[str] = None,
    keyword: Optional[str] = None,
    shopName: Optional[str] = None,
    detailsScraped: Optional[str] = None,
    minQuality: Optional[int] = None,
    page: int = 1,
    limit: int = 50,
    sort: str = "-createdAt",
):
    db = request.app.state.db
    filt = {}

    if platform:
        filt["platform"] = platform
    if category:
        filt["categoryId"] = category
    if keyword:
        filt["searchKeyword"] = {"$regex": re.escape(keyword), "$options": "i"}
    if detailsScraped is not None:
        filt["detailsScraped"] = detailsScraped.lower() == "true"
    if minQuality is not None:
        filt["extractionQuality"] = {"$gte": minQuality}
    if shopName:
        escaped = re.escape(shopName)
        filt["$or"] = [
            {"shopName": {"$regex": escaped, "$options": "i"}},
            {"shopInfo.shopName": {"$regex": escaped, "$options": "i"}},
        ]

    skip = (page - 1) * limit
    sort_key = sort.lstrip("-")
    sort_dir = -1 if sort.startswith("-") else 1

    import asyncio
    products, total = await asyncio.gather(
        db.products.find(filt, {"_id": 0})
            .sort(sort_key, sort_dir)
            .limit(limit)
            .skip(skip)
            .to_list(length=limit),
        db.products.count_documents(filt),
    )

    return JSONResponse({
        "success": True,
        "data": {
            "products": clean(products),
            "pagination": {
                "page": page,
                "limit": limit,
                "total": total,
                "pages": -(-total // limit),
            },
        },
    })


@router.get("/stats/summary")
async def stats_summary(request: Request):
    import asyncio
    db = request.app.state.db

    total, by_platform, by_category, with_details = await asyncio.gather(
        db.products.count_documents({}),
        db.products.aggregate([
            {"$group": {"_id": "$platform", "count": {"$sum": 1}}}
        ]).to_list(length=None),
        db.products.aggregate([
            {"$group": {"_id": "$categoryName", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
            {"$limit": 10},
        ]).to_list(length=None),
        db.products.count_documents({"detailsScraped": True}),
    )

    return JSONResponse({
        "success": True,
        "data": {
            "totalProducts": total,
            "productsWithDetails": with_details,
            "detailsPercentage": f"{(with_details / total * 100):.2f}" if total else "0.00",
            "byPlatform": clean(by_platform),
            "topCategories": clean(by_category),
        },
    })


@router.get("/search/text")
async def text_search(
    request: Request,
    q: str = Query(..., description="Search query"),
    page: int = 1,
    limit: int = 20,
):
    db = request.app.state.db
    skip = (page - 1) * limit
    products = await db.products.find(
        {"$text": {"$search": q}},
        {"_id": 0, "score": {"$meta": "textScore"}}
    ).sort([("score", {"$meta": "textScore"})]).skip(skip).limit(limit).to_list(length=limit)

    return JSONResponse({"success": True, "data": {"products": clean(products), "count": len(products)}})


@router.get("/export/json")
async def export_products_json(
    request: Request,
    platform: Optional[str] = None,
    keyword: Optional[str] = None,
    category: Optional[str] = None,
    detailsScraped: Optional[str] = None,
    minQuality: Optional[int] = None,
):
    """Export all matching products to a JSON file and return it for download."""
    db = request.app.state.db
    filt: dict = {}

    if platform:
        filt["platform"] = platform
    if keyword:
        filt["searchKeyword"] = {"$regex": re.escape(keyword), "$options": "i"}
    if category:
        filt["categoryName"] = {"$regex": re.escape(category), "$options": "i"}
    if detailsScraped is not None:
        filt["detailsScraped"] = detailsScraped.lower() == "true"
    if minQuality is not None:
        filt["extractionQuality"] = {"$gte": minQuality}

    products = await db.products.find(filt, {"_id": 0}).limit(_EXPORT_MAX).to_list(length=_EXPORT_MAX)

    if not products:
        raise HTTPException(404, "No products match the given filters")

    # Write to file in data/ directory
    export_dir = Path(__file__).parent.parent / "data"
    export_dir.mkdir(exist_ok=True)
    timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    filename = f"products_export_{timestamp}.json"
    filepath = export_dir / filename

    cleaned = clean(products)
    filepath.write_text(json.dumps(cleaned, indent=2, ensure_ascii=False), encoding="utf-8")

    return FileResponse(
        path=str(filepath),
        filename=filename,
        media_type="application/json",
    )


def _flatten_product(p: dict) -> dict:
    """Flatten nested product data into a single-level dict for CSV/Excel."""
    detail = p.get("detailedInfo") or {}
    shop = p.get("shopInfo") or detail.get("shopInfo") or {}
    quality = detail.get("dataQuality") or {}
    seller = shop.get("sellerInfo") or {}

    # Flatten variants: "口味: 辣味, 甜味; 净含量: 500g"
    variants = detail.get("variants") or {}
    variants_str = "; ".join(
        f"{k}: {', '.join(o['value'] if isinstance(o, dict) else str(o) for o in v)}"
        for k, v in variants.items()
        if isinstance(v, list)
    ) if isinstance(variants, dict) else str(variants)

    # Flatten specs: "品牌: X; 产地: Y"
    specs = detail.get("specifications") or {}
    specs_str = "; ".join(f"{k}: {v}" for k, v in specs.items()) if isinstance(specs, dict) else str(specs)

    return {
        "itemId":            p.get("itemId"),
        "title":             p.get("title"),
        "fullTitle":         detail.get("fullTitle"),
        "price":             p.get("price"),
        "priceDetail":       detail.get("price"),
        "priceUsd":          detail.get("priceUsd"),
        "originalPrice":     detail.get("originalPrice"),
        "originalPriceUsd":  detail.get("originalPriceUsd"),
        "platform":          p.get("platform"),
        "link":              p.get("link"),
        "image":             p.get("image"),
        "shopName":          shop.get("shopName") or p.get("shopName"),
        "shopRating":        shop.get("shopRating"),
        "shopLocation":      shop.get("shopLocation") or p.get("location"),
        "shopAge":           shop.get("shopAge"),
        "salesCount":        p.get("salesCount"),
        "salesVolume":       detail.get("salesVolume"),
        "rating":            detail.get("rating"),
        "reviewCount":       detail.get("reviewCount"),
        "brand":             detail.get("brand"),
        "description":       detail.get("fullDescription"),
        "specifications":    specs_str,
        "variants":          variants_str,
        "guarantees":        "; ".join(detail.get("guarantees") or []),
        "additionalImages":  ", ".join(detail.get("additionalImages") or []),
        "searchKeyword":     p.get("searchKeyword"),
        "categoryName":      p.get("categoryName"),
        "language":          p.get("language"),
        "detailsScraped":    p.get("detailsScraped"),
        "extractionQuality": p.get("extractionQuality"),
        "completeness":      quality.get("completeness"),
        "createdAt":         p.get("createdAt"),
        "updatedAt":         p.get("updatedAt"),
    }


_EXPORT_MAX = 10000  # Hard limit for export to prevent OOM


def _build_export_filter(
    platform: Optional[str], keyword: Optional[str],
    category: Optional[str], detailsScraped: Optional[str],
    minQuality: Optional[int],
) -> dict:
    filt: dict = {}
    if platform:
        filt["platform"] = platform
    if keyword:
        filt["searchKeyword"] = {"$regex": re.escape(keyword), "$options": "i"}
    if category:
        filt["categoryName"] = {"$regex": re.escape(category), "$options": "i"}
    if detailsScraped is not None:
        filt["detailsScraped"] = detailsScraped.lower() == "true"
    if minQuality is not None:
        filt["extractionQuality"] = {"$gte": minQuality}
    return filt


@router.get("/export/csv")
async def export_products_csv(
    request: Request,
    platform: Optional[str] = None,
    keyword: Optional[str] = None,
    category: Optional[str] = None,
    detailsScraped: Optional[str] = None,
    minQuality: Optional[int] = None,
):
    """Export matching products as a downloadable CSV file."""
    db = request.app.state.db
    filt = _build_export_filter(platform, keyword, category, detailsScraped, minQuality)
    products = await db.products.find(filt, {"_id": 0}).limit(_EXPORT_MAX).to_list(length=_EXPORT_MAX)

    if not products:
        raise HTTPException(404, "No products match the given filters")

    rows = [_flatten_product(p) for p in clean(products)]
    headers = list(rows[0].keys())

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    writer.writerows(rows)
    output.seek(0)

    timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=products_export_{timestamp}.csv"},
    )


@router.get("/export/excel")
async def export_products_excel(
    request: Request,
    platform: Optional[str] = None,
    keyword: Optional[str] = None,
    category: Optional[str] = None,
    detailsScraped: Optional[str] = None,
    minQuality: Optional[int] = None,
):
    """Export matching products as a downloadable Excel file."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    db = request.app.state.db
    filt = _build_export_filter(platform, keyword, category, detailsScraped, minQuality)
    products = await db.products.find(filt, {"_id": 0}).limit(_EXPORT_MAX).to_list(length=_EXPORT_MAX)

    if not products:
        raise HTTPException(404, "No products match the given filters")

    rows = [_flatten_product(p) for p in clean(products)]
    headers = list(rows[0].keys())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    # Header row with bold style
    from openpyxl.styles import Font
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(headers, 1):
            val = row.get(key)
            # Convert non-string types for Excel compatibility
            if isinstance(val, (dict, list)):
                val = str(val)
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-adjust column widths (capped at 50)
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=products_export_{timestamp}.xlsx"},
    )


@router.get("/{item_id}")
async def get_product(request: Request, item_id: str):
    db = request.app.state.db
    product = await db.products.find_one({"itemId": item_id}, {"_id": 0})
    if not product:
        raise HTTPException(404, "Product not found")
    return JSONResponse({"success": True, "data": clean(product)})


@router.delete("/{item_id}")
async def delete_product(request: Request, item_id: str):
    db = request.app.state.db
    result = await db.products.delete_one({"itemId": item_id})
    if result.deleted_count == 0:
        raise HTTPException(404, "Product not found")
    return JSONResponse({"success": True, "message": "Product deleted successfully"})